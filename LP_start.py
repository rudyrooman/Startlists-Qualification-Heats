# -*- coding: utf-8 -*-
"""
@author: Rudy Rooman rudy.rooman@gmail.com
"""

# -*- coding: utf-8 -*-
# from __future__ import print_function
from __future__ import unicode_literals
from ortools.linear_solver import pywraplp
from timeit import default_timer
from xlsxwriter import Workbook
from random import sample
from openpyxl import load_workbook
from pandas import DataFrame
from sys import exit
from itertools import combinations
from numpy import array_split


class Runner:
    def __init__(self):
        self.ID = 0
        self.FED = ''
        self.Surname = ''
        self.Firstname = ''
        self.StartGrp = 0
        self.RankingPoints = 0
        self.Rank = 0
        self.Heat = 0
        self.Time = 0
        runners.append(self)

    def __str__(self):
        return str(self.Firstname + self.Surname)


class Nation:
    def __init__(self):
        self.FED = ''
        self.runners = []
        self.count = 0
        self.groupcount = [0, 0, 0, 0]
        nations.append(self)

    def countrunners(self, _runners):
        # runners per nation
        self.runners = [_r for _r in _runners if self.FED == _r.FED]
        self.count = len(self.runners)
        # runners per nation in a startgroup
        for _i in range(4):
            self.groupcount[_i] = len([_r for _r in self.runners if _r.StartGrp == _i])

    def __str__(self):
        return str(self.FED)


def find_heats_time(_runners, _heats, _nations, _z):
    # solver = pywraplp.Solver('SolveAssignmentProblemMIP', pywraplp.Solver.CBC_MIXED_INTEGER_PROGRAMMING)
    solver = pywraplp.Solver('SolveAssignmentProblemMIP', pywraplp.Solver.SCIP_MIXED_INTEGER_PROGRAMMING)
    # solver = pywraplp.Solver('SolveAssignmentLinearProgramming', pywraplp.Solver.GLOP_LINEAR_PROGRAMMING)

    # define runners per heat using numpy array_split ( 10 runners over 3 heats = 4,3,3 )
    runners_per_heat = [len(it) for it in array_split(range(len(_runners)), _heats)]

    # startingblocks given by teammanagers ( 0 = no preference, 1 = early, 2 mid section, 3 =late)
    # count runners per starting block
    starting_blocks = [len([_r for _r in runners if _r.StartGrp == sb]) for sb in range(4)]
    # add runners without startblock preference to startgroup with the least athletes
    index = starting_blocks.index(min(starting_blocks[1:]))
    starting_blocks[index] += starting_blocks[0]
    starting_blocks = starting_blocks[1:]

    print()
    print('...Optimizer is running...Please wait...')
    print()

    # [ START create variables ]
    match = {}
    for _r in _runners:
        for _h in range(1, _heats + 1):
            for _t in range(runners_per_heat[_h - 1]):
                match[_r, _h, _t] = solver.BoolVar('match[%s,%s,%s]' % (_r, _h, _t))

    # [ END create variables ]

    # [ START Constraints ]

    # Each runner is assigned to exactly 1 heat / time combination
    for _r in _runners:
        solver.Add(solver.Sum([match[_r, _h, _t] for _h in range(1, _heats + 1)
                               for _t in range(runners_per_heat[_h - 1])]) == 1)

    # each heat / time combination is assigned to exactly 1 runner
    for _h in range(1, _heats + 1):
        for _t in range(runners_per_heat[_h - 1]):
            solver.Add(solver.Sum([match[_r, _h, _t] for _r in _runners]) == 1)

    # balance number of runners from 1 country to a heat
    for _n in _nations:
        for _h in range(1, _heats + 1):
            solver.Add(solver.Sum(
                [match[_r, _h, _t] for _t in range(runners_per_heat[_h - 1]) for _r in _n.runners])
                       <= (1 + (_n.count - 1) // _heats))
            solver.Add(solver.Sum(
                [match[_r, _h, _t] for _t in range(runners_per_heat[_h - 1]) for _r in _n.runners])
                       >= (_n.count // _heats))

    # spreading runners with rank 1,2,3 over different heats
    # spreading runners with rank 4,5,6 over different heats
    # ...
    # remove last runners to ensure all group have size = nr of heats
    _runners2 = _runners[:(len(_runners) // _heats) * _heats]
    # divide in groups
    for group in [list(it) for it in array_split(_runners2, len(_runners) // _heats)]:
        for _h in range(1, _heats + 1):
            solver.Add(solver.Sum([match[_r, _h, _t] for _r in group for _t in range(runners_per_heat[_h - 1])]) == 1)

    # consecutive times not from same nation
    for _n in _nations:
        # all combinations of 2 different runners for that nation
        for _r1, _r2 in combinations(_n.runners, 2):
            for _h in range(1, _heats + 1):
                for _t in range(runners_per_heat[_h - 1] - 1):
                    solver.Add(match[_r1, _h, _t] + match[_r2, _h, _t + 1] <= 1)

    # comply with startgroup requests
    # a parameter _z for relaxation is available
    for _r in _runners:
        if _r.StartGrp > 1:
            solver.Add(solver.Sum([match[_r, _h, _t] * _t for _h in range(1, _heats + 1)
                                   for _t in range(runners_per_heat[_h - 1])]) >=
                       ((sum(starting_blocks[0:_r.StartGrp - 1]) - 1) // _heats - _z))
        if _r.StartGrp < _heats and _r.StartGrp != 0:
            solver.Add(solver.Sum([match[_r, _h, _t] * _t for _h in range(1, _heats + 1)
                                   for _t in range(runners_per_heat[_h - 1])]) <=
                       ((sum(starting_blocks[0:_r.StartGrp]) - 1) // _heats + _z))

    # choose random runners and fix to specific heats and time
    random_runners = sample(_runners, _heats)
    for _h, _r in enumerate(random_runners, start=1):
        solver.Add(solver.Sum([match[_r, _h, _t] for _t in range(runners_per_heat[_h - 1])]) == 1)

    # [ END Constraints ]

    sol = solver.Solve()

    if sol == solver.OPTIMAL:
        # resultaten tonen
        print()
        if _z == 0:
            print('Starting times: Optimal solution found')
        elif _z > 0:
            print('Starting times: Solution found with correction factor = %i' % _z)

        print()
        print('runners per heat: %s' % runners_per_heat)
        print()
        print('runners per starting block: %s' % starting_blocks)
        print()
        print('Following runners are fixed to a heat to ensure random startlists.')
        for _c, _r in enumerate(random_runners, start=1):
            print('%s to heat %i.' % (_r, _c))
        print()
        for _h in range(1, _heats + 1):
            for _r in _runners:
                for _t in range(runners_per_heat[_h - 1]):
                    if match[_r, _h, _t].solution_value() == 1:
                        _r.Heat = _h
                        _r.Time = _t
    if sol == solver.OPTIMAL:
        txt = 'OPTIMAL'
    else:
        txt = 'FAILED'

    return solver, txt


# ###program starts here### #

print('###################################################')
print('##### STARTLISTS TOOL with LINEAR PROGRAMMING #####')
print('###################################################')
print()

heats = 3

# list of instances
runners = []
nations = []

start_time = default_timer()

# read entries
wb = load_workbook(filename='LP_start_entries.xlsx')
sheet1 = wb.active

# read entered runners data file
teller = 5
while sheet1.cell(row=teller, column=1).value:
    # read a row and create runner instances
    runner = Runner()
    runner.ID = sheet1.cell(row=teller, column=1).value
    runner.FED = sheet1.cell(row=teller, column=2).value
    runner.Surname = sheet1.cell(row=teller, column=3).value
    runner.Firstname = sheet1.cell(row=teller, column=4).value
    runner.StartGrp = sheet1.cell(row=teller, column=5).value
    runner.RankingPoints = sheet1.cell(row=teller, column=6).value
    teller += 1

# remove doubles from list of countrynames and create country instances
for country in list(dict.fromkeys([r.FED for r in runners])):
    n = Nation()
    n.FED = country

# count runners per nation and startgrps
for n in nations:
    n.countrunners(runners)

print('Startgroup Validation')
print()
startgrouperror = False
# 3 startgroups being 1 = early, 2 = middle, 3 = late

# check unexpected start group entries
unexpected = [r.StartGrp for r in runners if r.StartGrp not in [0, 1, 2, 3]]
if len(unexpected) > 0:
    startgrouperror = True
    print('Among startgroup entries we found following unexpected data:', end=' ')
    print(*unexpected, sep=", ")

# check on grp 0 runners
grp0runners = [r for r in runners if r.StartGrp == 0]
if len(grp0runners) > 0:
    startgrouperror = True
    print('Currently you have %i runners without startgroup or startgroup 0.' % len(grp0runners))

# check on startgroups with too many runners
for n in nations:
    for i in range(1, 4):
        if n.groupcount[i] > 1 + (n.count - 1) // 3:
            print('Too many runners from %s in startgroup %i ' % (n.FED, i))
            startgrouperror = True

# give user the choice to proceed
if startgrouperror:
    print()
    print('We strongly recommend to correct startgroups before proceeding.')
    print('Typ "s" and enter to stop program or "p" to proceed:', end=' ')
    answer = ''
    while answer not in ['p', 's']:
        answer = input()
    if answer == 's':
        exit()
else:
    print('Startgroup Validation completed without comments.')
    print()

# sort participants based on ranking
runners = sorted(runners, key=lambda x: x.RankingPoints, reverse=True)

print('We have %i entries.' % len(runners))

# save rank of sorted runners
for teller, r in enumerate(runners, start=1):
    r.Rank = teller

# optimise
z = 0
while True:
    solution, optimal_result = find_heats_time(runners, heats, nations, z)
    if optimal_result == 'OPTIMAL':
        break
    z += 1

print('Making startlists.xlxsx')

# sort participants based on heat, starttimes
runners = sorted(runners, key=lambda x: (x.Heat, x.Time))

# print en export to Excel
workbook = Workbook('startlists.xlsx')
# engineer data from Clicksoftware
startlist_sheet = workbook.add_worksheet('startlist')
row = 0
col = 0
startlist_sheet.write(row, col, 'Heat')
startlist_sheet.write(row, col + 1, 'Time')
startlist_sheet.write(row, col + 2, 'Firstname')
startlist_sheet.write(row, col + 3, 'Surname')
startlist_sheet.write(row, col + 4, 'FED')
startlist_sheet.write(row, col + 5, 'Rank')
startlist_sheet.write(row, col + 6, 'RankingPoints')
startlist_sheet.write(row, col + 7, 'ID')
startlist_sheet.write(row, col + 8, 'StartGrp')
row = 1
for r in runners:
    print(r.Heat, r.Time, r, r.FED, r.Rank, r.ID, sep=";")
    startlist_sheet.write(row, col, r.Heat)
    startlist_sheet.write(row, col + 1, r.Time)
    startlist_sheet.write(row, col + 2, r.Firstname)
    startlist_sheet.write(row, col + 3, r.Surname)
    startlist_sheet.write(row, col + 4, r.FED)
    startlist_sheet.write(row, col + 5, r.Rank)
    startlist_sheet.write(row, col + 6, r.RankingPoints)
    startlist_sheet.write(row, col + 7, r.ID)
    startlist_sheet.write(row, col + 8, r.StartGrp)
    row += 1

workbook.close()

elapsed = default_timer() - start_time

print()
print('Calculation time: %s seconds.' % round(elapsed, 3))

#
# Verification
#

dfver = DataFrame([vars(r) for r in runners])

print("******************")
print("Number of runners per federation & startgroup")
print(dfver.groupby(['FED', 'StartGrp']).count()[['ID']])

print("******************")
print("Number of runners per federation & heat")
print(dfver.groupby(['FED', 'Heat']).count()[['ID']])

print("******************")
print("Number of runners per federation & heat - min, max and diff")

runnersperheat = dfver.groupby(['FED', 'Heat']).count()[['ID']]

t = runnersperheat.assign(ID=runnersperheat.ID.abs()) \
    .groupby('FED') \
    .ID.agg([('Min', 'min'), ('Max', 'max')]) \
    .add_prefix('Count')

t['Diff'] = t.apply(lambda rij: rij.CountMax - rij.CountMin, axis=1)
print(t)

print("******************")
print("Average ranking per heat")
print(dfver.groupby(['Heat']).mean(numeric_only=True)[['RankingPoints']])

